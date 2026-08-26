from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = PROJECT_ROOT / "outputs" / "day02" / "TeacherOS_Day02_Research_Workbook.xlsx"

PRIVATE_PATTERNS = (
    re.compile(r"\bhttps?://", re.IGNORECASE),
    re.compile(r"\bt\.me/", re.IGNORECASE),
    re.compile(r"\b[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}\b"),
    re.compile(r"(?<!\d)(?:\+?98|0)?9\d{9}(?!\d)"),
)


@dataclass
class GateResult:
    status: str
    metrics: dict[str, int]
    blockers: list[str]
    warnings: list[str]


def _text(value: Any) -> str:
    return str(value or "").strip()


def _int(value: Any, fallback: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def _score(value: Any) -> bool:
    return 1 <= _int(value) <= 5


def _rows_by_header(sheet: Any, header_row: int = 4) -> list[dict[str, Any]]:
    headers = [_text(cell.value) for cell in sheet[header_row]]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        if any(_text(value) for value in values):
            rows.append(row)
    return rows


def _decision_values(sheet: Any) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for row in range(5, 17):
        key = _text(sheet.cell(row=row, column=1).value)
        if key:
            values[key] = sheet.cell(row=row, column=2).value
    return values


def _contains_private_locator(value: Any) -> bool:
    text = _text(value)
    return bool(text and any(pattern.search(text) for pattern in PRIVATE_PATTERNS))


def evaluate_workbook(path: Path = DEFAULT_WORKBOOK) -> GateResult:
    blockers: list[str] = []
    warnings: list[str] = []
    if not path.is_file():
        return GateResult(
            status="CLOSED",
            metrics={},
            blockers=[f"Research workbook is missing: {path}"],
            warnings=[],
        )

    workbook = load_workbook(path, data_only=False, read_only=False)
    required_sheets = {
        "Dashboard", "Recruitment", "Interviews", "Job Scores", "Hypotheses",
        "Feature Map", "Safety Requirements", "Scoring", "Decision",
    }
    missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
    if missing_sheets:
        blockers.append("Missing workbook sheets: " + ", ".join(missing_sheets))
        return GateResult("CLOSED", {}, blockers, warnings)

    recruitment = _rows_by_header(workbook["Recruitment"])
    interviews = _rows_by_header(workbook["Interviews"])
    feature_rows = _rows_by_header(workbook["Feature Map"])
    safety_rows = _rows_by_header(workbook["Safety Requirements"])
    decision = _decision_values(workbook["Decision"])

    eligible_participants: set[str] = set()
    confirmed_pilot: set[str] = set()
    for row in recruitment:
        participant_id = _text(row.get("participant_id"))
        lessons = _int(row.get("recurring_lessons_per_week"), -1)
        telegram_daily = _text(row.get("telegram_daily")).lower()
        eligible = lessons >= 2 and telegram_daily == "yes"
        if eligible and participant_id:
            eligible_participants.add(participant_id)
        if _text(row.get("pilot_status")).lower() == "confirmed":
            if not eligible:
                blockers.append(f"{participant_id or 'Unknown participant'} is confirmed for pilot but is not eligible.")
            elif participant_id:
                confirmed_pilot.add(participant_id)
        for key in ("source_channel", "deidentified_notes"):
            if _contains_private_locator(row.get(key)):
                blockers.append(f"Private contact-like data detected in Recruitment {participant_id or '?'} field {key}.")

    completed: dict[str, dict[str, Any]] = {}
    completed_participants: set[str] = set()
    problem_counts: Counter[str] = Counter()
    artifact_total = 0
    for row in interviews:
        if _text(row.get("status")).lower() != "completed":
            continue
        interview_id = _text(row.get("interview_id"))
        participant_id = _text(row.get("participant_id"))
        if not interview_id or interview_id in completed:
            blockers.append(f"Completed interview has a missing or duplicate ID: {interview_id or '(blank)' }.")
            continue
        completed[interview_id] = row
        if participant_id in completed_participants:
            blockers.append(f"Participant {participant_id} is counted in more than one completed interview.")
        completed_participants.add(participant_id)

        eligible = participant_id in eligible_participants and _text(row.get("eligible")).lower() == "yes"
        if not eligible:
            blockers.append(f"Completed interview {interview_id} is not tied to an eligible recruitment row.")
            continue

        required_text = (
            "last_real_task", "artifact_ids", "problem_code", "pain_observed",
            "current_workaround", "consequence", "desired_outcome", "proof_of_success",
            "trust_concern",
        )
        for field in required_text:
            if not _text(row.get(field)):
                blockers.append(f"Completed interview {interview_id} is missing {field}.")
        if _int(row.get("artifact_count")) < 1:
            blockers.append(f"Completed interview {interview_id} has no artifact walkthrough.")
        else:
            artifact_total += _int(row.get("artifact_count"))
        if _int(row.get("minutes_spent")) < 1:
            blockers.append(f"Completed interview {interview_id} is missing a positive task-time estimate.")
        for field in (
            "frequency_1_5", "time_loss_1_5", "consequence_1_5",
            "upload_willingness_1_5", "pay_willingness_1_5",
        ):
            if not _score(row.get(field)):
                blockers.append(f"Completed interview {interview_id} has an invalid {field} score.")
        problem_code = _text(row.get("problem_code"))
        if problem_code:
            problem_counts[problem_code] += 1
        for field, value in row.items():
            if _contains_private_locator(value):
                blockers.append(f"Private contact-like data detected in interview {interview_id} field {field}.")

    eligible_completed = [
        row for row in completed.values()
        if _text(row.get("participant_id")) in eligible_participants
        and _text(row.get("eligible")).lower() == "yes"
    ]
    if len(eligible_completed) < 5:
        blockers.append(f"Need at least 5 completed eligible interviews; found {len(eligible_completed)}.")
    if not 10 <= len(confirmed_pilot) <= 15:
        blockers.append(f"Need 10-15 confirmed eligible pilot recruits; found {len(confirmed_pilot)}.")

    repeated_problems = {code for code, count in problem_counts.items() if count >= 3}
    if not repeated_problems:
        blockers.append("No problem is independently repeated by at least 3 eligible teachers.")

    research_status = _text(decision.get("research_status")).lower()
    if research_status != "validated":
        blockers.append("Decision research_status must remain hypothesis until evidence is complete, then be set to validated.")
    for field in ("target_segment", "anti_segment", "product_promise"):
        if not _text(decision.get(field)):
            blockers.append(f"Decision record is missing {field}.")
    top_problem_codes = [
        _text(decision.get(f"top_problem_{index}")) for index in range(1, 4)
    ]
    top_problem_codes = [code for code in top_problem_codes if code]
    if not top_problem_codes:
        blockers.append("Decision record must select at least one repeated problem and no more than three.")
    if len(top_problem_codes) != len(set(top_problem_codes)):
        blockers.append("Decision record contains duplicate top problem codes.")
    for code in top_problem_codes:
        if code not in repeated_problems:
            blockers.append(f"Top problem {code} is not repeated in at least 3 eligible completed interviews.")

    safety_ids = {
        _text(row.get("requirement_id")) for row in safety_rows
        if _text(row.get("status")).lower() == "approved"
    }
    feature_ids = [_text(row.get("feature_id")) for row in feature_rows]
    if len(feature_rows) != 90 or len(set(feature_ids)) != 90:
        blockers.append(f"Feature map must contain 90 unique Day 6-23 features; found {len(feature_rows)} rows and {len(set(feature_ids))} unique IDs.")
    day_counts = Counter(_int(row.get("day")) for row in feature_rows)
    for day in range(6, 24):
        if day_counts[day] != 5:
            blockers.append(f"Feature map Day {day} must contain 5 bullets; found {day_counts[day]}.")

    approved_features = 0
    deferred_features = 0
    for row in feature_rows:
        feature_id = _text(row.get("feature_id")) or "unknown feature"
        mapping_type = _text(row.get("mapping_type")).lower()
        mapping_id = _text(row.get("mapping_id"))
        if mapping_type == "safety_reliability":
            if mapping_id not in safety_ids:
                blockers.append(f"{feature_id} references an unapproved safety requirement {mapping_id or '(blank)'}.")
            else:
                approved_features += 1
        elif mapping_type == "witnessed_task":
            evidence_ids = {
                item.strip() for item in re.split(r"[;,]", _text(row.get("evidence_ids"))) if item.strip()
            }
            problem_code = _text(row.get("problem_code"))
            if not evidence_ids or not evidence_ids.issubset(completed.keys()):
                blockers.append(f"{feature_id} must reference only completed interview IDs.")
            elif problem_code not in repeated_problems:
                blockers.append(f"{feature_id} maps to {problem_code or '(blank)'}, which lacks repeated evidence.")
            else:
                approved_features += 1
        elif mapping_type == "defer":
            if not _text(row.get("decision_notes")):
                blockers.append(f"{feature_id} is deferred without a reason.")
            else:
                deferred_features += 1
                approved_features += 1
        else:
            blockers.append(f"{feature_id} remains an unvalidated hypothesis; map it to witnessed evidence or defer it.")

    metrics = {
        "eligible_recruits": len(eligible_participants),
        "completed_eligible_interviews": len(eligible_completed),
        "confirmed_pilot_recruits": len(confirmed_pilot),
        "artifact_walkthroughs": artifact_total,
        "repeated_problem_codes": len(repeated_problems),
        "approved_or_deferred_features": approved_features,
        "deferred_features": deferred_features,
        "feature_total": len(feature_rows),
    }
    return GateResult(
        status="OPEN" if not blockers else "CLOSED",
        metrics=metrics,
        blockers=blockers,
        warnings=warnings,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate the TeacherOS Day 2 evidence gate.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON")
    args = parser.parse_args()
    result = evaluate_workbook(args.workbook.resolve())
    if args.json:
        print(json.dumps(asdict(result), indent=2))
    else:
        print(f"DAY 2 GATE: {result.status}")
        for name, value in result.metrics.items():
            print(f"- {name}: {value}")
        if result.blockers:
            print("\nBlockers:")
            for blocker in result.blockers:
                print(f"- {blocker}")
        for warning in result.warnings:
            print(f"WARNING: {warning}")
    raise SystemExit(0 if result.status == "OPEN" else 1)


if __name__ == "__main__":
    main()
